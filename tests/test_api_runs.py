def test_post_run_creates_run_and_dispatches_orchestration(api_client, auth_header, monkeypatch):
    calls = []
    monkeypatch.setattr(
        "api.routers.runs.orchestrate_run.delay", lambda run_id, source_ids: calls.append((run_id, source_ids))
    )

    response = api_client.post("/runs", json={"source_ids": [1, 2]}, headers=auth_header)

    assert response.status_code == 202
    body = response.json()
    assert body["status"] == "pending"
    assert body["triggered_by"] == "manual"
    assert calls == [(body["id"], [1, 2])]


def test_get_run_returns_404_for_unknown_id(api_client, auth_header):
    response = api_client.get("/runs/999999", headers=auth_header)
    assert response.status_code == 404


def test_cancel_run_sets_cancel_requested(api_client, auth_header, monkeypatch, db_session):
    from core.db import repository

    monkeypatch.setattr("api.routers.runs.orchestrate_run.delay", lambda *a, **k: None)

    create_response = api_client.post("/runs", json={}, headers=auth_header)
    run_id = create_response.json()["id"]

    cancel_response = api_client.post(f"/runs/{run_id}/cancel", headers=auth_header)
    assert cancel_response.status_code == 200
    assert cancel_response.json()["cancel_requested"] is True


def test_get_runs_respects_limit_and_offset(api_client, auth_header, monkeypatch):
    monkeypatch.setattr("api.routers.runs.orchestrate_run.delay", lambda *a, **k: None)

    for _ in range(3):
        api_client.post("/runs", json={}, headers=auth_header)

    response = api_client.get("/runs?limit=2", headers=auth_header)
    assert response.status_code == 200
    assert len(response.json()) == 2

    response = api_client.get("/runs?limit=2&offset=2", headers=auth_header)
    assert response.status_code == 200
    assert len(response.json()) == 1


def test_get_runs_rejects_out_of_range_limit_and_offset(api_client, auth_header):
    assert api_client.get("/runs?offset=-1", headers=auth_header).status_code == 422
    assert api_client.get("/runs?limit=0", headers=auth_header).status_code == 422
    assert api_client.get("/runs?limit=1000000", headers=auth_header).status_code == 422


def test_retry_failed_dispatches_retry_task_for_failed_sources_and_sets_run_running(
    api_client, auth_header, monkeypatch, db_session
):
    from core.db import repository

    monkeypatch.setattr("api.routers.runs.orchestrate_run.delay", lambda *a, **k: None)
    calls = []
    monkeypatch.setattr(
        "api.routers.runs.retry_failed_run_sources_task.delay",
        lambda run_id, run_source_ids: calls.append((run_id, run_source_ids)),
    )

    repository.create_source_family(db_session, key="samai", display_name="SAMAI")
    ok_source = repository.create_source(db_session, family_key="samai", name="Tribunal OK", family_params={})
    bad_source = repository.create_source(db_session, family_key="samai", name="Tribunal Fallido", family_params={})
    run = repository.create_run(db_session, triggered_by="manual", fini=None, ffin=None)
    ok_run_source = repository.create_run_source(db_session, run_id=run.id, source_id=ok_source.id)
    bad_run_source = repository.create_run_source(db_session, run_id=run.id, source_id=bad_source.id)
    repository.set_run_source_status(db_session, ok_run_source.id, "completed")
    repository.set_run_source_status(db_session, bad_run_source.id, "failed", error_message="boom")
    repository.set_run_status(db_session, run.id, "failed")

    response = api_client.post(f"/runs/{run.id}/retry-failed", headers=auth_header)

    assert response.status_code == 200
    assert response.json()["status"] == "running"
    assert calls == [(run.id, [bad_run_source.id])]


def test_retry_failed_returns_400_when_no_source_failed(api_client, auth_header, monkeypatch, db_session):
    from core.db import repository

    monkeypatch.setattr("api.routers.runs.orchestrate_run.delay", lambda *a, **k: None)
    repository.create_source_family(db_session, key="samai", display_name="SAMAI")
    source = repository.create_source(db_session, family_key="samai", name="Tribunal OK", family_params={})
    run = repository.create_run(db_session, triggered_by="manual", fini=None, ffin=None)
    run_source = repository.create_run_source(db_session, run_id=run.id, source_id=source.id)
    repository.set_run_source_status(db_session, run_source.id, "completed")
    repository.set_run_status(db_session, run.id, "completed")

    response = api_client.post(f"/runs/{run.id}/retry-failed", headers=auth_header)

    assert response.status_code == 400


def test_retry_failed_returns_404_for_unknown_run(api_client, auth_header):
    response = api_client.post("/runs/999999/retry-failed", headers=auth_header)
    assert response.status_code == 404


def test_get_run_sources_reports_docs_updated(api_client, auth_header, monkeypatch, db_session):
    from core.db import repository

    monkeypatch.setattr("api.routers.runs.orchestrate_run.delay", lambda *a, **k: None)
    repository.create_source_family(db_session, key="constitucional", display_name="Corte Constitucional")
    source = repository.create_source(db_session, family_key="constitucional", name="Corte Constitucional", family_params={})
    run = repository.create_run(db_session, triggered_by="manual", fini=None, ffin=None)
    run_source = repository.create_run_source(db_session, run_id=run.id, source_id=source.id)
    repository.set_run_source_status(db_session, run_source.id, "completed", docs_new=1, docs_updated=3, docs_errors=0)

    response = api_client.get(f"/runs/{run.id}/sources", headers=auth_header)

    assert response.status_code == 200
    assert response.json()[0]["docs_updated"] == 3
    assert response.json()[0]["source_name"] == "Corte Constitucional"


def test_get_run_report_returns_404_for_unknown_run(api_client, auth_header):
    response = api_client.get("/runs/999999/report.xlsx", headers=auth_header)
    assert response.status_code == 404


def test_get_run_report_has_header_only_sheets_for_a_run_with_no_documents(api_client, auth_header, monkeypatch, db_session):
    from io import BytesIO

    from openpyxl import load_workbook

    from core.db import repository

    monkeypatch.setattr("api.routers.runs.orchestrate_run.delay", lambda *a, **k: None)
    run = repository.create_run(db_session, triggered_by="manual", fini=None, ffin=None)

    response = api_client.get(f"/runs/{run.id}/report.xlsx", headers=auth_header)

    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content))
    assert wb.sheetnames == ["Resumen", "Documentos nuevos", "Documentos actualizados", "Errores"]
    for sheet_name in wb.sheetnames:
        assert wb[sheet_name].max_row == 1


def test_get_run_report_lists_new_updated_and_error_documents_by_run(api_client, auth_header, monkeypatch, db_session):
    from io import BytesIO

    from openpyxl import load_workbook

    from core.db import repository

    monkeypatch.setattr("api.routers.runs.orchestrate_run.delay", lambda *a, **k: None)
    repository.create_source_family(db_session, key="constitucional", display_name="Corte Constitucional")
    source = repository.create_source(db_session, family_key="constitucional", name="Corte Constitucional", family_params={})
    run = repository.create_run(db_session, triggered_by="manual", fini=None, ffin=None)
    run_source = repository.create_run_source(db_session, run_id=run.id, source_id=source.id)

    repository.insert_document(
        db_session,
        doc_id="report-new-doc",
        source_id=source.id,
        run_source_id=run_source.id,
        title="T-100/26",
        storage_bucket="iurisync-test",
        storage_key="Corte Constitucional/2026-01-01/Sentencia/T-100-26.rtf",
        source_url="https://example.org/T-100-26.rtf",
    )

    existing = repository.insert_document(
        db_session,
        doc_id="report-updated-doc",
        source_id=source.id,
        title="A. 200/26",
        storage_bucket="iurisync-test",
        storage_key="Corte Constitucional/2026-01-02/Auto/A.200-26.rtf",
        source_url="https://example.org/A.200-26.rtf",
    )
    repository.archive_and_replace_document(
        db_session,
        existing.id,
        run_source_id=run_source.id,
        storage_bucket="iurisync-test",
        storage_key="Corte Constitucional/2026-01-02/Auto/A.200-26-republicado.rtf",
    )

    repository.add_run_error(
        db_session, run_source.id, "Timeout descargando el archivo", context={"title": "T-999/26", "url": "https://example.org/T-999-26.rtf"}
    )
    repository.set_run_source_status(db_session, run_source.id, "completed", docs_new=1, docs_updated=1, docs_errors=1)

    response = api_client.get(f"/runs/{run.id}/report.xlsx", headers=auth_header)

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert f'informe_run_{run.id}.xlsx' in response.headers["content-disposition"]

    wb = load_workbook(BytesIO(response.content))
    assert wb.sheetnames == ["Resumen", "Documentos nuevos", "Documentos actualizados", "Errores"]

    resumen = wb["Resumen"]
    assert [cell.value for cell in resumen[2]] == ["Corte Constitucional", "completed", 1, 1, 1]

    nuevos = wb["Documentos nuevos"]
    assert nuevos.cell(row=2, column=1).value == "Corte Constitucional"
    assert nuevos.cell(row=2, column=2).value == "T-100/26"

    actualizados = wb["Documentos actualizados"]
    assert actualizados.cell(row=2, column=1).value == "Corte Constitucional"
    assert actualizados.cell(row=2, column=2).value == "A. 200/26"

    errores = wb["Errores"]
    assert errores.cell(row=2, column=1).value == "Corte Constitucional"
    assert errores.cell(row=2, column=2).value == "T-999/26"
    assert errores.cell(row=2, column=4).value == "Timeout descargando el archivo"


def test_get_run_report_only_lists_documents_updated_by_that_specific_run(api_client, auth_header, monkeypatch, db_session):
    """Regression guard for the run_source_id wiring in archive_and_replace_document:
    updating the same document twice, in two different runs, must attribute each
    republication to the run that actually performed it — not both to the latest one."""
    from core.db import repository

    monkeypatch.setattr("api.routers.runs.orchestrate_run.delay", lambda *a, **k: None)
    repository.create_source_family(db_session, key="constitucional", display_name="Corte Constitucional")
    source = repository.create_source(db_session, family_key="constitucional", name="Corte Constitucional", family_params={})

    run_a = repository.create_run(db_session, triggered_by="manual", fini=None, ffin=None)
    run_source_a = repository.create_run_source(db_session, run_id=run_a.id, source_id=source.id)
    run_b = repository.create_run(db_session, triggered_by="manual", fini=None, ffin=None)
    run_source_b = repository.create_run_source(db_session, run_id=run_b.id, source_id=source.id)

    document = repository.insert_document(
        db_session,
        doc_id="report-twice-updated-doc",
        source_id=source.id,
        title="A. 300/26",
        storage_bucket="iurisync-test",
        storage_key="Corte Constitucional/2026-01-03/Auto/A.300-26.rtf",
    )
    repository.archive_and_replace_document(
        db_session,
        document.id,
        run_source_id=run_source_a.id,
        storage_bucket="iurisync-test",
        storage_key="Corte Constitucional/2026-01-03/Auto/A.300-26-v2.rtf",
    )
    repository.archive_and_replace_document(
        db_session,
        document.id,
        run_source_id=run_source_b.id,
        storage_bucket="iurisync-test",
        storage_key="Corte Constitucional/2026-01-03/Auto/A.300-26-v3.rtf",
    )

    updated_in_a = repository.list_updated_documents_for_run(db_session, run_a.id)
    updated_in_b = repository.list_updated_documents_for_run(db_session, run_b.id)

    assert len(updated_in_a) == 1
    assert len(updated_in_b) == 1
    assert updated_in_a[0]["title"] == "A. 300/26"
    assert updated_in_b[0]["title"] == "A. 300/26"
