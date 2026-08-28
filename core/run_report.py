import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font


def _write_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for column_cells in ws.columns:
        lengths = [len(str(cell.value)) for cell in column_cells if cell.value is not None]
        width = max(lengths, default=0) + 2
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width, 10), 60)


def _naive(value):
    # openpyxl rejects timezone-aware datetimes ("Excel does not support timezones
    # in datetimes") — every timestamp we store is tz-aware (DateTime(timezone=True)),
    # so this has to run on every one of them before it reaches a cell.
    return value.replace(tzinfo=None) if value is not None else None


def build_run_report_workbook(
    run_sources: list,
    new_documents: list[dict],
    updated_documents: list[dict],
    errors: list[dict],
) -> bytes:
    """Builds the run report as .xlsx bytes: one summary sheet mirroring the
    per-source counts already shown on the run detail page, plus one sheet per
    category listing exactly which documents fell into it."""
    wb = Workbook()

    resumen = wb.active
    resumen.title = "Resumen"
    _write_sheet(
        resumen,
        ["Fuente", "Estado", "Docs nuevos", "Actualizados", "Docs con error"],
        [[rs.source_name, rs.status, rs.docs_new, rs.docs_updated, rs.docs_errors] for rs in run_sources],
    )

    nuevos = wb.create_sheet("Documentos nuevos")
    _write_sheet(
        nuevos,
        [
            "Fuente",
            "Título",
            "Tipo",
            "Sección",
            "Especialidad",
            "Magistrado",
            "Fecha publicación",
            "Fecha providencia",
            "URL",
            "Descargado el",
        ],
        [
            [
                d["source_name"],
                d["title"],
                d["tipo"],
                d["seccion"],
                d["especialidad"],
                d["magistrado"],
                d["f_public"],
                d["f_providencia"],
                d["source_url"],
                _naive(d["downloaded_at"]),
            ]
            for d in new_documents
        ],
    )

    actualizados = wb.create_sheet("Documentos actualizados")
    _write_sheet(
        actualizados,
        ["Fuente", "Título", "URL", "Actualizado el"],
        [[d["source_name"], d["title"], d["source_url"], _naive(d["updated_at"])] for d in updated_documents],
    )

    errores = wb.create_sheet("Errores")
    _write_sheet(
        errores,
        ["Fuente", "Documento", "URL", "Mensaje", "Ocurrido el"],
        [[e["source_name"], e["title"], e["url"], e["message"], _naive(e["occurred_at"])] for e in errors],
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
